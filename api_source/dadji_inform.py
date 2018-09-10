from openpyxl import Workbook
import openpyxl
from urllib.request import urlopen
import xml.etree.ElementTree as ET
import time

# 단지 정보가져오기
def xl_header(nws) :
    nws.cell(row=1, column=1).value = '건축면적(㎡)'
    nws.cell(row=1, column=2).value = '부속건축물면적(㎡)'
    nws.cell(row=1, column=3).value = '부속건축물수'
    nws.cell(row=1, column=4).value = '건폐율(%)'
    nws.cell(row=1, column=5).value = '새주소법정동코드'
    nws.cell(row=1, column=6).value = '건물명'
    nws.cell(row=1, column=7).value = '블록'
    nws.cell(row=1, column=8).value = '번'
    nws.cell(row=1, column=9).value = '지'
    nws.cell(row=1, column=10).value = '외필지수'
    nws.cell(row=1, column=11).value = '생성일자'
    nws.cell(row=1, column=12).value = 'EPI점수'
    nws.cell(row=1, column=13).value = '에너지절감율'
    nws.cell(row=1, column=14).value = '기타용도'
    nws.cell(row=1, column=15).value = '가구수(가구)'
    nws.cell(row=1, column=16).value = '부속건축물수'
    nws.cell(row=1, column=17).value = '친환경건축물인증점수'
    nws.cell(row=1, column=18).value = '친환경건축물인증점수'
    nws.cell(row=1, column=19).value = '세대수(세대)'
    nws.cell(row=1, column=20).value = '호수(호)'
    nws.cell(row=1, column=21).value = '옥내자주식면적(㎡)'
    nws.cell(row=1, column=22).value = '옥내자주식대수(대)'
    nws.cell(row=1, column=23).value = '옥내기계식면적(㎡)'
    nws.cell(row=1, column=24).value = '옥내기계식대수(대)'
    nws.cell(row=1, column=25).value = '지능형건축물인증점수'
    nws.cell(row=1, column=26).value = '지능형건축물등급'
    nws.cell(row=1, column=27).value = '로트'
    nws.cell(row=1, column=28).value = '주건축물수'
    nws.cell(row=1, column=29).value = '주용도코드'
    nws.cell(row=1, column=30).value = '주용도코드명'
    nws.cell(row=1, column=31).value = '관리건축물대장PK'
    nws.cell(row=1, column=32).value = '새주소법정동코드'
    nws.cell(row=1, column=33).value = '새주소본번'
    nws.cell(row=1, column=34).value = '새주소도로코드'
    nws.cell(row=1, column=35).value = '새주소부번'
    nws.cell(row=1, column=36).value = '새주소지상지하코드'
    nws.cell(row=1, column=37).value = '신구대장구분코드'
    nws.cell(row=1, column=38).value = '신구대장구분코드명'
    nws.cell(row=1, column=39).value = '도로명대지위치'
    nws.cell(row=1, column=40).value = '옥외자주식면적(㎡)'
    nws.cell(row=1, column=41).value = '옥외자주식대수(대)'
    nws.cell(row=1, column=42).value = '옥외기계식면적(㎡)'
    nws.cell(row=1, column=43).value = '옥외기계식대수(대)'
    nws.cell(row=1, column=44).value = '대지면적(㎡)'
    nws.cell(row=1, column=45).value = '대지구분코드'
    nws.cell(row=1, column=46).value = '대지위치'
    nws.cell(row=1, column=47).value = '허가일'
    nws.cell(row=1, column=48).value = '허가번호구분코드'
    nws.cell(row=1, column=49).value = '허가번호구분코드명'
    nws.cell(row=1, column=50).value = '허가번호기관코드'
    nws.cell(row=1, column=51).value = '허가번호기관코드명'
    nws.cell(row=1, column=52).value = '허가번호년'
    nws.cell(row=1, column=53).value = '대장구분코드'
    nws.cell(row=1, column=54).value = '대장구분코드명'
    nws.cell(row=1, column=55).value = '대장종류코드'
    nws.cell(row=1, column=56).value = '대장종류코드명'
    nws.cell(row=1, column=57).value = '순번'
    nws.cell(row=1, column=58).value = '시군구코드'
    nws.cell(row=1, column=59).value = '특수지명'
    nws.cell(row=1, column=60).value = '착공일'
    nws.cell(row=1, column=61).value = '연면적(㎡)'
    nws.cell(row=1, column=62).value = '총주차수'
    nws.cell(row=1, column=63).value = '사용승인일'
    nws.cell(row=1, column=64).value = '용적률(%)'
    nws.cell(row=1, column=65).value = '용적률산정연면적(㎡)'
    nws.cell(row=1, column=66).value = '주소(위경도측정용)'

total_count = 0;
part_seq = 9;
try :
    wb = openpyxl.load_workbook('C:/Users/kyungwon/Documents/pythonExcel/region_code_exist.xlsx');
    ws = wb.active;
    nwb = openpyxl.Workbook();
    nws = nwb.active;
    xl_header(nws);


    service_key = #data.go.kr 서비스 코드입력


    write_row = 2;
    for r in ws.rows:
        c_num = 50;
        region_code = str(r[0].value)  # 행 인덱스
        sigungu_cd  = region_code[:5]
        bjdong_cd   = region_code[-5:]
        # 지역코드 한건당 아파트 코드 api를 읽어 아파트 코드 worksheet만들기

        if  int(region_code) >= 0:
            API_HOST1 =  'http://apis.data.go.kr/1611000/BldRgstService/getBrTitleInfo?serviceKey='+service_key+'&sigunguCd=' + sigungu_cd + '&bjdongCd=' + bjdong_cd + '&platGbCd=0&numOfRows=10&pageNo=1&startPage=1'
            file1 = urlopen(API_HOST1)
            tree1 = ET.ElementTree(file=file1)
            root1 = tree1.getroot()
            if len(root1[1][3].text) != 0 :
                total_count = int(root1[1][3].text)
            else :
                total_count = 0;
            if total_count > c_num :
                total_page = int(total_count / c_num) + 1 ;
            else:
                total_page = 1;
            time.sleep(1);
            print('-----------------'+str(region_code)+'-----------------')
            for i in range(1,total_page + 1):
                API_HOST = 'http://apis.data.go.kr/1611000/BldRgstService/getBrTitleInfo?serviceKey=' + service_key + '&sigunguCd=' + sigungu_cd + '&bjdongCd=' + bjdong_cd + '&platGbCd=0&numOfRows=' + str(c_num) + '&pageNo='+str(i)+'&startPage=1'
                file = urlopen(API_HOST)
                tree = ET.ElementTree(file=file)
                root = tree.getroot()
                for data in root.iter("item"):
                    if write_row > 30000:
                        nwb.save("C:/Users/kyungwon/Documents/pythonExcel/danji_inform" + str(part_seq) + ".xlsx");
                        nwb.close();
                        nwb = openpyxl.Workbook();
                        nws = nwb.active;
                        xl_header(nws)
                        write_row = 2;
                        part_seq = part_seq + 1;

                    archArea = data.findtext("archArea")
                    atchBldArea = data.findtext("atchBldArea")
                    atchBldCnt = data.findtext("atchBldCnt")
                    bcRat = data.findtext("bcRat")
                    bjdongCd = data.findtext("bjdongCd")
                    bldNm = data.findtext("bldNm")
                    block = data.findtext("block")
                    bun = data.findtext("bun")
                    ji = data.findtext("ji")
                    bylotCnt = data.findtext("bylotCnt")
                    crtnDay = data.findtext("crtnDay")
                    engrEpi = data.findtext("engrEpi")
                    engrRat = data.findtext("engrRat")
                    etcPurps = data.findtext("etcPurps")
                    fmlyCnt = data.findtext("fmlyCnt")
                    atchBldCnt = data.findtext("atchBldCnt")
                    gnBldCert = data.findtext("gnBldCert")
                    gnBldGrade = data.findtext("gnBldGrade")
                    hhldCnt = data.findtext("hhldCnt")
                    hoCnt = data.findtext("hoCnt")
                    indrAutoArea = data.findtext("indrAutoArea")
                    indrAutoUtcnt = data.findtext("indrAutoUtcnt")
                    indrMechArea = data.findtext("indrMechArea")
                    indrMechUtcnt = data.findtext("indrMechUtcnt")
                    itgBldCert = data.findtext("itgBldCert")
                    itgBldGrade = data.findtext("itgBldGrade")
                    lot = data.findtext("lot")
                    mainBldCnt = data.findtext("mainBldCnt")
                    mainPurpsCd = data.findtext("mainPurpsCd")
                    mainPurpsCdNm = data.findtext("mainPurpsCdNm")
                    mgmBldrgstPk = data.findtext("mgmBldrgstPk")
                    naBjdongCd = data.findtext("naBjdongCd")
                    naMainBun = data.findtext("naMainBun")
                    naRoadCd = data.findtext("naRoadCd")
                    naSubBun = data.findtext("naSubBun")
                    naUgrndCd = data.findtext("naUgrndCd")
                    newOldRegstrGbCd = data.findtext("newOldRegstrGbCd")
                    newOldRegstrGbCdNm = data.findtext("newOldRegstrGbCdNm")
                    newPlatPlc = data.findtext("newPlatPlc")
                    oudrAutoArea = data.findtext("oudrAutoArea")
                    oudrAutoUtcnt = data.findtext("oudrAutoUtcnt")
                    oudrMechArea = data.findtext("oudrMechArea")
                    oudrMechUtcnt = data.findtext("oudrMechUtcnt")
                    platArea = data.findtext("platArea")
                    platGbCd = data.findtext("platGbCd")
                    platPlc = data.findtext("platPlc")
                    pmsDay = data.findtext("pmsDay")
                    pmsnoGbCd = data.findtext("pmsnoGbCd")
                    pmsnoGbCdNm = data.findtext("pmsnoGbCdNm")
                    pmsnoKikCd = data.findtext("pmsnoKikCd")
                    pmsnoKikCdNm = data.findtext("pmsnoKikCdNm")
                    pmsnoYear = data.findtext("pmsnoYear")
                    regstrGbCd = data.findtext("regstrGbCd")
                    regstrGbCdNm = data.findtext("regstrGbCdNm")
                    regstrKindCd = data.findtext("regstrKindCd")
                    regstrKindCdNm = data.findtext("regstrKindCdNm")
                    rnum = data.findtext("rnum")
                    sigunguCd = data.findtext("sigunguCd")
                    splotNm = data.findtext("splotNm")
                    stcnsDay = data.findtext("stcnsDay")
                    totArea = data.findtext("totArea")
                    totPkngCnt = data.findtext("totPkngCnt")
                    useAprDay = data.findtext("useAprDay")
                    vlRat = data.findtext("vlRat")
                    vlRatEstmTotArea = data.findtext("vlRatEstmTotArea")


                    address_make =  platPlc.replace("번지", "");

                    if(mainPurpsCd == '02000') :
                        print(platPlc,region_code,i);
                        nws.cell(row=write_row, column=1).value = archArea
                        nws.cell(row=write_row, column=2).value = atchBldArea
                        nws.cell(row=write_row, column=3).value = atchBldCnt

                        nws.cell(row=write_row, column=4).value =bcRat
                        nws.cell(row=write_row, column=5).value =bjdongCd
                        nws.cell(row=write_row, column=6).value =bldNm
                        nws.cell(row=write_row, column=7).value = block
                        nws.cell(row=write_row, column=8).value =bun
                        nws.cell(row=write_row, column=9).value =ji
                        nws.cell(row=write_row, column=10).value =bylotCnt
                        nws.cell(row=write_row, column=11).value =crtnDay
                        nws.cell(row=write_row, column=12).value =engrEpi
                        nws.cell(row=write_row, column=13).value =engrRat
                        nws.cell(row=write_row, column=14).value =etcPurps
                        nws.cell(row=write_row, column=15).value =fmlyCnt
                        nws.cell(row=write_row, column=16).value =atchBldCnt
                        nws.cell(row=write_row, column=17).value =gnBldCert
                        nws.cell(row=write_row, column=18).value =gnBldGrade
                        nws.cell(row=write_row, column=19).value =hhldCnt
                        nws.cell(row=write_row, column=20).value =hoCnt
                        nws.cell(row=write_row, column=21).value =indrAutoArea
                        nws.cell(row=write_row, column=22).value =indrAutoUtcnt
                        nws.cell(row=write_row, column=23).value =indrMechArea
                        nws.cell(row=write_row, column=24).value =indrMechUtcnt
                        nws.cell(row=write_row, column=25).value =itgBldCert
                        nws.cell(row=write_row, column=26).value =itgBldGrade
                        nws.cell(row=write_row, column=27).value =lot
                        nws.cell(row=write_row, column=28).value =mainBldCnt
                        nws.cell(row=write_row, column=29).value =mainPurpsCd
                        nws.cell(row=write_row, column=30).value =mainPurpsCdNm
                        nws.cell(row=write_row, column=31).value =mgmBldrgstPk
                        nws.cell(row=write_row, column=32).value =naBjdongCd
                        nws.cell(row=write_row, column=33).value =naMainBun
                        nws.cell(row=write_row, column=34).value =naRoadCd
                        nws.cell(row=write_row, column=35).value =naSubBun

                        nws.cell(row=write_row, column=36).value =naUgrndCd
                        nws.cell(row=write_row, column=37).value =newOldRegstrGbCd
                        nws.cell(row=write_row, column=38).value =newOldRegstrGbCdNm
                        nws.cell(row=write_row, column=39).value =newPlatPlc
                        nws.cell(row=write_row, column=40).value =oudrAutoArea
                        nws.cell(row=write_row, column=41).value =oudrAutoUtcnt
                        nws.cell(row=write_row, column=42).value =oudrMechArea
                        nws.cell(row=write_row, column=43).value =oudrMechUtcnt
                        nws.cell(row=write_row, column=44).value =platArea
                        nws.cell(row=write_row, column=45).value =platGbCd
                        nws.cell(row=write_row, column=46).value =platPlc
                        nws.cell(row=write_row, column=47).value =pmsDay
                        nws.cell(row=write_row, column=48).value =pmsnoGbCd
                        nws.cell(row=write_row, column=49).value =pmsnoGbCdNm

                        nws.cell(row=write_row, column=50).value =pmsnoKikCd
                        nws.cell(row=write_row, column=51).value =pmsnoKikCdNm
                        nws.cell(row=write_row, column=52).value =pmsnoYear
                        nws.cell(row=write_row, column=53).value =regstrGbCd
                        nws.cell(row=write_row, column=54).value =regstrGbCdNm
                        nws.cell(row=write_row, column=55).value =regstrKindCd
                        nws.cell(row=write_row, column=56).value =regstrKindCdNm
                        nws.cell(row=write_row, column=57).value =rnum
                        nws.cell(row=write_row, column=58).value =sigunguCd
                        nws.cell(row=write_row, column=59).value =splotNm
                        nws.cell(row=write_row, column=60).value =stcnsDay

                        nws.cell(row=write_row, column=61).value =totArea
                        nws.cell(row=write_row, column=62).value =totPkngCnt
                        nws.cell(row=write_row, column=63).value =useAprDay
                        nws.cell(row=write_row, column=64).value =vlRat
                        nws.cell(row=write_row, column=65).value =vlRatEstmTotArea

                        nws.cell(row=write_row, column=66).value = address_make
                        nws.cell(row=write_row, column=67).value = region_code

                        write_row = write_row + 1;
            time.sleep(0.3);

except Exception as ex:
    print('에러발생',ex);
    print(API_HOST1)
    nwb.save("C:/Users/kyungwon/Documents/pythonExcel/danji_inform"+str(part_seq)+".xlsx");
    nwb.close();
    wb.close();
# 엑셀 파일 저장
nwb.save("C:/Users/kyungwon/Documents/pythonExcel/danji_inform"+str(part_seq)+".xlsx");
nwb.close();
wb.close();