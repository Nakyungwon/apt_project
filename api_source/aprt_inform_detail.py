from openpyxl import Workbook
import openpyxl
from urllib.request import urlopen
import xml.etree.ElementTree as ET
import time

total_count = 0;
part_seq = 1;
p_year = '' #<--가져오고자하는 년도 입력
service_key = '' #<--data.go.kr 서비스키 입력
c_num = '' #<--한번에 가져올 건데이터 입력
try :
    #시군구코드 엑셀파일 불러오기
    wb = openpyxl.load_workbook('C:/Users/kyungwon/Documents/pythonExcel/region_code_detail.xlsx');
    ws = wb.active;
    nwb = openpyxl.Workbook();
    nws = nwb.active;

    #엑셀로우
    write_row = 1;
    for r in ws.rows:
        region_code = str(r[0].value)  # 행 인덱스
        if int(region_code) >= 0 :

            #최초 접근해서 해당조건을 만족하는 아파트 건수 확인
            API_HOST1 = 'http://openapi.molit.go.kr/OpenAPI_ToolInstallPackage/service/rest/RTMSOBJSvc/getRTMSDataSvcAptTradeDev?serviceKey='+service_key+'&pageNo=1&startPage=1&numOfRows=10&pageSize=10&LAWD_CD='+region_code+'&DEAL_YMD='+p_year+'%25'
            tree1 = ET.ElementTree(file=urlopen(API_HOST1))
            root1 = tree1.getroot()
            if root1[0][0].text != '00':
                print('api 접근 에러발생',root1[0][1].text);
                raise Exception
            if len(root1[1][3].text) != 0 :
                total_count = int(root1[1][3].text)
            else :
                total_count = 0;
            if total_count > c_num :
                total_page = int(total_count / c_num) + 1 ;
            else:
                total_page = 1;

            for i in range(1,total_page + 1):
                API_HOST = 'http://openapi.molit.go.kr/OpenAPI_ToolInstallPackage/service/rest/RTMSOBJSvc/getRTMSDataSvcAptTradeDev?serviceKey='+service_key+'&pageNo='+str(i)+'&startPage=1&numOfRows='+str(c_num)+'&pageSize=10&LAWD_CD='+region_code+'&DEAL_YMD='+p_year+'%25'
                tree = ET.ElementTree(file=urlopen(API_HOST))
                root = tree.getroot()
                if root[0][0].text != '00':
                    print('api 접근 에러발생', root[0][1].text);
                    raise Exception
                for data in root.iter("item"):
                    if write_row > 100000:
                        nwb.save("C:/Users/kyungwon/Documents/pythonExcel/apt_inform_detail" + str(part_seq) + ".xlsx");
                        nwb.close();
                        nwb = openpyxl.Workbook();
                        nws = nwb.active;
                        write_row = 1;
                        part_seq = part_seq + 1;
                    #건축년도 1
                    build_year = data.findtext("건축년도");
                    # 거래금액 (만원)2
                    price = data.findtext("거래금액");
                    #년3
                    year = data.findtext("년");
                    #월4
                    month = data.findtext("월");
                    #일5
                    day = data.findtext("일");
                    #법정동6
                    dong = data.findtext("법정동");
                    #지번7
                    ji = data.findtext("지번");
                    #아파트이름8
                    apt_name = data.findtext("아파트");
                    #층9
                    floor = data.findtext("층");
                    # 전용면적10
                    measure = data.findtext("전용면적");
                    #주소 조합10
                    address = str(r[1].value)+' '+str(dong)+' '+str(ji);
                    #일련번호11
                    srial = data.findtext("일련번호");
                    #시군구코드+법정동코드12
                    area_code = str(data.findtext("법정동시군구코드"))+str(data.findtext("법정동읍면동코드"));

                    print(build_year,apt_name,address,area_code, i)
                    nws.cell(row=write_row, column=1).value = build_year;
                    nws.cell(row=write_row, column=2).value = price;
                    nws.cell(row=write_row, column=3).value = year;
                    nws.cell(row=write_row, column=4).value = month;
                    nws.cell(row=write_row, column=5).value = day;
                    nws.cell(row=write_row, column=6).value = dong;
                    nws.cell(row=write_row, column=7).value = ji;
                    nws.cell(row=write_row, column=8).value = apt_name;
                    nws.cell(row=write_row, column=9).value = floor;
                    nws.cell(row=write_row, column=10).value = measure;
                    nws.cell(row=write_row, column=11).value = address;
                    nws.cell(row=write_row, column=12).value = srial;
                    nws.cell(row=write_row, column=13).value = area_code;
                    nws.cell(row=write_row, column=14).value = i;
                    write_row = write_row + 1;
            time.sleep(1);

except Exception as ex:
    print('에러발생',total_count,ex);
    print(API_HOST1);
    nwb.save("C:/Users/kyungwon/Documents/pythonExcel/apt_inform_detail"+str(part_seq)+".xlsx");
    nwb.close();
    wb.close();
# 엑셀 파일 저장
nwb.save("C:/Users/kyungwon/Documents/pythonExcel/apt_inform_detail"+str(part_seq)+".xlsx");
nwb.close();
wb.close();